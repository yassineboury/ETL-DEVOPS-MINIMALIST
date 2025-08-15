"""
Formatters Excel - Styles et formatage pour les exports Excel
Module séparé pour réduire la complexité cognitive
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelStyleFormatter:
    """Classe responsable du formatage et des styles Excel"""
    
    # Constantes de style
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    CONTENT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
    BORDER_STYLE = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def apply_header_style(self, worksheet, max_col: int):
        """
        Applique le style d'en-tête aux cellules de la première ligne
        
        Args:
            worksheet: Feuille Excel
            max_col: Nombre maximum de colonnes
        """
        for col_num in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_STYLE

    def apply_content_alignment(self, worksheet):
        """
        Applique l'alignement aux cellules de contenu
        
        Args:
            worksheet: Feuille Excel
        """
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value is not None:
                    cell.alignment = self.CONTENT_ALIGNMENT

    def setup_auto_filter_and_freeze(self, worksheet):
        """
        Configure le filtre automatique et le gel des en-têtes
        
        Args:
            worksheet: Feuille Excel
        """
        if worksheet.max_row > 0 and worksheet.max_column > 0:
            max_col_letter = get_column_letter(worksheet.max_column)
            worksheet.auto_filter.ref = f"A1:{max_col_letter}{worksheet.max_row}"
            worksheet.freeze_panes = "A2"


class ExcelColumnAdjuster:
    """Classe responsable de l'ajustement automatique des colonnes"""
    
    # Configuration des largeurs minimales par type de colonne
    MIN_WIDTHS = {
        'id': 8, 'nom': 15, 'email': 25, 'username': 15, 'url': 30,
        'date': 18, 'etat': 12, 'admin': 8, 'type': 12,
        'langage': 15, 'namespace': 20, 'complet': 25
    }
    
    MAX_COLUMN_WIDTH = 60
    DEFAULT_MIN_WIDTH = 12
    SPACING_MARGIN = 3
    
    def auto_adjust_columns(self, worksheet):
        """
        Ajuste automatiquement la largeur des colonnes
        
        Args:
            worksheet: Feuille Excel
        """
        for column in worksheet.columns:
            column_name = get_column_letter(column[0].column)
            header = str(column[0].value or "").lower()
            
            min_width = self._get_min_width_for_column(header)
            max_length = self._calculate_max_column_length(column)
            final_width = self._calculate_final_width(max_length, min_width)
            
            worksheet.column_dimensions[column_name].width = final_width

    def _get_min_width_for_column(self, header: str) -> int:
        """Détermine la largeur minimale basée sur le type de colonne"""
        for key, width in self.MIN_WIDTHS.items():
            if key in header:
                return width
        return self.DEFAULT_MIN_WIDTH

    def _calculate_max_column_length(self, column) -> int:
        """Calcule la longueur maximale nécessaire pour une colonne"""
        max_length = 0
        
        for cell in column:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                cell_length = len(cell_value)
                
                if cell_length > 0:
                    cell_length += self._get_bonus_length(cell_value)
                
                max_length = max(max_length, cell_length)
            except (AttributeError, ValueError, TypeError):
                continue
        
        return max_length

    def _get_bonus_length(self, cell_value: str) -> int:
        """Calcule la longueur bonus basée sur le type de contenu"""
        if '@' in cell_value or 'http' in cell_value.lower():
            return 2
        elif '/' in cell_value and len(cell_value) > 8:
            return 1
        return 0

    def _calculate_final_width(self, max_length: int, min_width: int) -> int:
        """Calcule la largeur finale de la colonne"""
        return min(max(max_length + self.SPACING_MARGIN, min_width), self.MAX_COLUMN_WIDTH)
