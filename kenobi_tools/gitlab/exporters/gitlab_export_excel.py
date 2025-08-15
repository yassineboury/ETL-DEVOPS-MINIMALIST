"""
Exporteur Excel pour GitLab
Module pour exporter les données GitLab vers Excel avec formatage
"""
from pathlib import Path
from typing import Any, Dict, Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Constantes
METRIC_COLUMN_NAME = 'Métrique'


class GitLabExcelExporter:
    """Classe pour exporter les données GitLab vers Excel avec formatage professionnel"""

    # Constantes pour les colonnes répétées
    DATE_CREATION_HEADER = 'Date Creation'
    ID_PROJET_HEADER = 'id Projet'

    def __init__(self, export_dir: Optional[Path] = None):
        """
        Initialise l'exporteur

        Args:
                        # Formatage basique
                worksheet = writer.sheets.get('Gitlab Groups')
                if worksheet is not None:
                    self._apply_header_style(worksheet, len(df_renamed.columns))
                    self._auto_adjust_columns(worksheet)

                    # Appliquer l'alignement à gauche pour le contenu
                    self._apply_content_alignment(worksheet)

                    # Ajouter un filtre automatique sur les en-têtes
                    from openpyxl.utils import get_column_letter
                    max_col_letter = get_column_letter(worksheet.max_column)
                    worksheet.auto_filter.ref = f"A1:{max_col_letter}{worksheet.max_row}"

                    worksheet.freeze_panes = "A2"

            print(f"✅ Fichier groupes Excel créé: {file_path}")
            print(f"📊 {len(df_renamed)} groupes exportés")pertoire d'export (défaut: exports/gitlab/)
        """
        if export_dir is None:
            # Trouver le dossier racine du projet (remonter 4 niveaux depuis kenobi_tools/gitlab/exporters/)
            current_dir = Path(__file__).parent.parent.parent.parent
            self.export_dir = current_dir / "exports" / "gitlab"
        else:
            self.export_dir = Path(export_dir)

        # Créer le dossier d'export s'il n'existe pas
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def _apply_header_style(self, worksheet, max_col: int):
        """
        Applique le style aux en-têtes

        Args:
            worksheet: Feuille Excel
            max_col: Nombre maximum de colonnes
        """
        # Style pour les en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="left", vertical="center")

        # Bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Appliquer le style aux en-têtes
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _auto_adjust_columns(self, worksheet):
        """
        Ajuste automatiquement la largeur des colonnes de façon optimale

        Args:
            worksheet: Feuille Excel
        """
        for column in worksheet.columns:
            column_name = column[0].column_letter
            header = str(column[0].value).lower() if column[0].value else ""

            min_width = self._get_min_width_for_column(header)
            max_length = self._calculate_max_column_length(column)
            final_width = self._calculate_final_width(max_length, min_width)

            worksheet.column_dimensions[column_name].width = final_width

    def _get_min_width_for_column(self, header: str) -> int:
        """Détermine la largeur minimale basée sur le type de colonne"""
        min_widths = {
            'id': 8, 'nom': 15, 'email': 25, 'username': 15, 'url': 30,
            'date': 18, 'etat': 12, 'admin': 8, 'type': 12,
            'langage': 15, 'namespace': 20, 'complet': 25
        }

        for key, width in min_widths.items():
            if key in header:
                return width
        return 12  # Valeur par défaut

    def _calculate_max_column_length(self, column) -> int:
        """Calcule la longueur maximale nécessaire pour une colonne"""
        max_length = 0

        for cell in column:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                cell_length = len(cell_value)

                if cell_length > 0:
                    cell_length += self._get_bonus_length(cell_value)

                max_length = max(max_length, cell_length)

            except (AttributeError, ValueError, TypeError):
                pass

        return max_length

    def _get_bonus_length(self, cell_value: str) -> int:
        """Calcule la longueur bonus basée sur le type de contenu"""
        # Bonus pour les URLs et emails
        if '@' in cell_value or 'http' in cell_value.lower():
            return 2
        # Bonus pour les dates
        elif '/' in cell_value and len(cell_value) > 8:
            return 1
        return 0

    def _calculate_final_width(self, max_length: int, min_width: int) -> int:
        """Calcule la largeur finale de la colonne"""
        # Minimum : basé sur le type de colonne
        # Ajout de marge : +3 pour l'espacement
        # Maximum : 60 pour éviter des colonnes trop larges
        return min(max(max_length + 3, min_width), 60)

    def _apply_content_alignment(self, worksheet):
        """
        Applique l'alignement à gauche pour tout le contenu des cellules (sauf en-têtes)

        Args:
            worksheet: Feuille Excel
        """
        from openpyxl.styles import Alignment

        # Alignement à gauche pour le contenu
        content_alignment = Alignment(horizontal="left", vertical="center")

        # Appliquer l'alignement à toutes les cellules de données
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        # Parcourir toutes les cellules sauf la première ligne (en-têtes)
        for row_num in range(2, max_row + 1):
            for col_num in range(1, max_col + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                if cell.value is not None:  # Seulement si la cellule a une valeur
                    cell.alignment = content_alignment

    def export_users(self, df_users: pd.DataFrame, filename: str = "gitlab_users.xlsx") -> str:
        """
        Exporte les utilisateurs vers Excel - VERSION SIMPLIFIÉE

        Args:
            df_users: DataFrame des utilisateurs
            filename: Nom du fichier

        Returns:
            Chemin complet du fichier créé
        """
        if df_users.empty:
            print("❌ Aucune donnée utilisateur à exporter")
            return ""

        try:
            file_path = self.export_dir / filename
            print(f"📁 Export vers: {file_path}")

            # Trier par date de création (plus récent en premier)
            if 'date_creation' in df_users.columns:
                # Convertir les dates pour un tri correct
                df_sorted = df_users.copy()

                # Créer une colonne temporaire pour le tri
                df_sorted['_temp_date'] = pd.to_datetime(
                    df_sorted['date_creation'],
                    format='%d/%m/%Y %H:%M:%S',
                    errors='coerce'
                )

                # Trier par date (plus récent en premier = descending)
                df_sorted = df_sorted.sort_values('_temp_date', ascending=False, na_position='last')

                # Supprimer la colonne temporaire
                df_sorted = df_sorted.drop('_temp_date', axis=1)

                print("📅 Utilisateurs triés par date de création (plus récent en premier)")
            else:
                df_sorted = df_users
                print("⚠️ Colonne 'date_creation' non trouvée, pas de tri")

            # Export direct avec une seule feuille
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Renommer les colonnes selon les spécifications
                column_mapping = {
                    'id_utilisateur': 'id Utilisateur',
                    'nom_utilisateur': 'Nom Utilisateur',
                    'email': 'Email',
                    'nom_complet': 'Nom Complet',
                    'admin': 'Admin',
                    'etat': 'Etat',
                    'type_utilisateur': 'Type Utilisateur',
                    'date_creation': self.DATE_CREATION_HEADER,
                    'date_validation': 'Date Validation',
                    'derniere_activite': 'Date Derniere Activite',
                    'derniere_connexion': 'Date Derniere Connexion'
                }

                # Appliquer le renommage des colonnes
                df_renamed = df_sorted.rename(columns=column_mapping)

                # UNE SEULE FEUILLE nommée "Gitlab Users"
                df_renamed.to_excel(writer, sheet_name='Gitlab Users', index=False)

                # Formatage et ajustement des colonnes
                workbook = writer.book
                worksheet = workbook['Gitlab Users']

                # Appliquer le formatage des en-têtes
                self._apply_header_style(worksheet, len(df_renamed.columns))

                # Ajuster automatiquement les largeurs de colonnes
                self._auto_adjust_columns(worksheet)

                # Appliquer l'alignement à gauche pour le contenu
                self._apply_content_alignment(worksheet)

                # Ajouter un filtre automatique sur les en-têtes
                from openpyxl.utils import get_column_letter
                max_col_letter = get_column_letter(worksheet.max_column)
                worksheet.auto_filter.ref = f"A1:{max_col_letter}{worksheet.max_row}"

                # Figer la première ligne
                worksheet.freeze_panes = "A2"

            print(f"✅ Fichier Excel créé: {file_path}")
            print(f"📊 {len(df_renamed)} utilisateurs exportés")

            return str(file_path)

        except Exception as e:
            print(f"❌ Erreur lors de l'export des utilisateurs: {e}")
            return ""

    def export_events(self, df_events: pd.DataFrame, filename: str = "gitlab_events.xlsx") -> str:
        """
        Exporte les événements vers Excel - VERSION ULTRA SIMPLE pour Power BI

        Args:
            df_events: DataFrame avec les données d'événements
            filename: Nom du fichier

        Returns:
            Chemin du fichier créé
        """
        try:
            file_path = self.export_dir / filename

            # Export direct sans limitation et sans formatage complexe
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # UNE SEULE FEUILLE : Données brutes pour Power BI
                df_events.to_excel(writer, sheet_name='Événements', index=False)

                # Formatage et ajustement des colonnes
                workbook = writer.book
                worksheet = workbook['Événements']

                # Appliquer le formatage des en-têtes
                self._apply_header_style(worksheet, len(df_events.columns))

                # Ajuster automatiquement les largeurs de colonnes
                self._auto_adjust_columns(worksheet)

                # Figer la première ligne
                worksheet.freeze_panes = "A2"

            print(f"✅ Export événements réussi: {file_path}")
            return str(file_path)

        except Exception as e:
            print(f"❌ Erreur lors de l'export des événements: {e}")
            return ""

    def _format_event_actions(self, worksheet, df_events: pd.DataFrame):
        """
        Applique un formatage conditionnel aux actions d'événements

        Args:
            worksheet: Feuille Excel
            df_events: DataFrame des événements
        """
        try:
            # Trouver la colonne nom_action
            action_col = None
            for idx, col in enumerate(df_events.columns, 1):
                if col == 'nom_action':
                    action_col = idx
                    break

            if action_col is None:
                return

            # Couleurs pour différents types d'actions
            action_colors = {
                'ouvert': 'C6EFCE',      # Vert clair
                'fermé': 'FFC7CE',       # Rouge clair
                'fusionné': 'FFEB9C',    # Jaune clair
                'poussé': 'DDEBF7',      # Bleu clair
                'commenté': 'E1D5E7',    # Violet clair
                'créé': 'D4EDDA',        # Vert très clair
                'mis à jour': 'F8D7DA'   # Rose clair
            }

            # Appliquer les couleurs ligne par ligne
            for row_idx in range(2, len(df_events) + 2):  # Commencer après l'en-tête
                cell = worksheet.cell(row=row_idx, column=action_col)
                if cell.value:
                    action_value = str(cell.value).lower()

                    for action, color in action_colors.items():
                        if action in action_value:
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                            break

        except Exception as e:
            print(f"⚠️ Erreur formatage actions: {e}")

    def _add_events_statistics(self, writer, df_events: pd.DataFrame, original_count: int | None = None):
        """
        Ajoute un onglet avec les statistiques des événements

        Args:
            writer: ExcelWriter
            df_events: DataFrame des événements
            original_count: Nombre original d'événements avant limitation (optionnel)
        """
        try:
            if df_events.empty:
                return

            stats_data = []

            self._add_general_statistics(stats_data, df_events, original_count)
            self._add_action_statistics(stats_data, df_events)
            self._add_target_statistics(stats_data, df_events)
            self._add_project_statistics(stats_data, df_events)
            self._add_author_statistics(stats_data, df_events)

            # Créer le DataFrame des statistiques
            stats_df = pd.DataFrame(stats_data, columns=[METRIC_COLUMN_NAME, 'Valeur'])

            # Exporter vers un onglet séparé
            stats_df.to_excel(writer, sheet_name='Statistiques', index=False)

            # Formater l'onglet statistiques
            workbook = writer.book
            stats_worksheet = workbook['Statistiques']
            self._apply_header_style(stats_worksheet, 2)

        except Exception as e:
            print(f"⚠️ Erreur lors de l'ajout des statistiques d'événements: {e}")

    def _add_general_statistics(self, stats_data: list, df_events: pd.DataFrame, original_count: int | None):
        """Ajoute les statistiques générales"""
        stats_data.append(['=== STATISTIQUES GÉNÉRALES ===', ''])
        if original_count and original_count != len(df_events):
            stats_data.append(['Nombre total d\'événements (original)', original_count])
            stats_data.append(['Nombre d\'événements dans ce fichier', len(df_events)])
            stats_data.append(['Note', f'Limité aux {len(df_events)} plus récents pour optimiser Excel'])
        else:
            stats_data.append(['Nombre total d\'événements', len(df_events)])

    def _add_action_statistics(self, stats_data: list, df_events: pd.DataFrame):
        """Ajoute les statistiques par type d'action"""
        if 'nom_action' in df_events.columns:
            action_counts = df_events['nom_action'].value_counts()
            stats_data.append(['', ''])
            stats_data.append(['=== PAR TYPE D\'ACTION ===', ''])
            for action, count in action_counts.items():
                stats_data.append([action, count])

    def _add_target_statistics(self, stats_data: list, df_events: pd.DataFrame):
        """Ajoute les statistiques par type de cible"""
        if 'type_cible' in df_events.columns:
            target_counts = df_events['type_cible'].value_counts()
            stats_data.append(['', ''])
            stats_data.append(['=== PAR TYPE DE CIBLE ===', ''])
            for target, count in target_counts.items():
                stats_data.append([target, count])

    def _add_project_statistics(self, stats_data: list, df_events: pd.DataFrame):
        """Ajoute les statistiques par projet (top 10)"""
        if 'id_projet' in df_events.columns:
            project_counts = df_events['id_projet'].value_counts().head(10)
            stats_data.append(['', ''])
            stats_data.append(['=== TOP 10 PROJETS ACTIFS ===', ''])
            for project, count in project_counts.items():
                stats_data.append([f'Projet {project}', count])

    def _add_author_statistics(self, stats_data: list, df_events: pd.DataFrame):
        """Ajoute les statistiques par auteur (top 10)"""
        if 'id_auteur' in df_events.columns:
            author_counts = df_events['id_auteur'].value_counts().head(10)
            stats_data.append(['', ''])
            stats_data.append(['=== TOP 10 UTILISATEURS ACTIFS ===', ''])
            for author, count in author_counts.items():
                stats_data.append([f'Utilisateur {author}', count])

    def export_merge_requests(
        self, df_mrs: pd.DataFrame, filename: str = "gitlab_merge_requests.xlsx"
    ) -> Optional[str]:
        """
        Exporte les Merge Requests vers Excel avec formatage et statistiques

        Args:
            df_mrs: DataFrame des Merge Requests
            filename: Nom du fichier

        Returns:
            Chemin complet du fichier créé
        """
        if df_mrs.empty:
            print("❌ Aucune donnée MR à exporter")
            return None

        try:
            output_path = self.export_dir / filename

            print(f"📝 Export de {len(df_mrs)} Merge Requests...")

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Onglet principal des MR
                df_mrs.to_excel(writer, sheet_name='Merge Requests', index=False)

                # Formater l'onglet principal
                workbook = writer.book
                worksheet = workbook['Merge Requests']

                self._apply_header_style(worksheet, len(df_mrs.columns))
                self._auto_adjust_columns(worksheet)

                # Figer la première ligne
                worksheet.freeze_panes = "A2"

                # Ajouter les statistiques
                self._add_mr_statistics(writer, df_mrs)

            print(f"✅ Export terminé: {output_path}")
            return str(output_path)

        except Exception as e:
            print(f"❌ Erreur export MR: {e}")
            return None

    def _add_general_mr_stats(self, stats_data: list, df_mrs: pd.DataFrame):
        """Ajoute les statistiques générales des MR"""
        stats_data.append(['=== STATISTIQUES MERGE REQUESTS ===', ''])
        stats_data.append(['Total MR', len(df_mrs)])

    def _add_state_stats(self, stats_data: list, df_mrs: pd.DataFrame):
        """Ajoute les statistiques par état"""
        if 'etat' in df_mrs.columns:
            state_counts = df_mrs['etat'].value_counts()
            stats_data.append(['', ''])
            stats_data.append(['=== RÉPARTITION PAR ÉTAT ===', ''])
            for state, count in state_counts.items():
                stats_data.append([f'État {state}', count])

    def _add_merge_status_stats(self, stats_data: list, df_mrs: pd.DataFrame):
        """Ajoute les statistiques par statut de fusion"""
        if 'statut_fusion' in df_mrs.columns:
            merge_status_counts = df_mrs['statut_fusion'].value_counts()
            stats_data.append(['', ''])
            stats_data.append(['=== STATUTS DE FUSION ===', ''])
            for status, count in merge_status_counts.items():
                stats_data.append([f'{status}', count])

    def _add_conflict_stats(self, stats_data: list, df_mrs: pd.DataFrame):
        """Ajoute les statistiques de conflits"""
        if 'conflits' in df_mrs.columns:
            conflicts_counts = df_mrs['conflits'].value_counts()
            stats_data.append(['', ''])
            stats_data.append(['=== CONFLITS ===', ''])
            for conflict, count in conflicts_counts.items():
                stats_data.append([f'Conflits: {conflict}', count])

    def _add_project_stats(self, stats_data: list, df_mrs: pd.DataFrame):
        """Ajoute les statistiques par projet (top 10)"""
        if 'id_projet' in df_mrs.columns:
            project_counts = df_mrs['id_projet'].value_counts().head(10)
            stats_data.append(['', ''])
            stats_data.append(['=== TOP 10 PROJETS AVEC MR ===', ''])
            for project, count in project_counts.items():
                stats_data.append([f'Projet {project}', count])

    def _add_author_stats(self, stats_data: list, df_mrs: pd.DataFrame):
        """Ajoute les statistiques par auteur (top 10)"""
        if 'id_auteur' in df_mrs.columns:
            author_counts = df_mrs['id_auteur'].value_counts().head(10)
            stats_data.append(['', ''])
            stats_data.append(['=== TOP 10 AUTEURS MR ===', ''])
            for author, count in author_counts.items():
                stats_data.append([f'Utilisateur {author}', count])

    def _add_mr_statistics(self, writer, df_mrs: pd.DataFrame):
        """Ajoute des statistiques sur les Merge Requests"""
        try:
            stats_data = []

            # Utiliser les fonctions helper pour réduire la complexité
            self._add_general_mr_stats(stats_data, df_mrs)
            self._add_state_stats(stats_data, df_mrs)
            self._add_merge_status_stats(stats_data, df_mrs)
            self._add_conflict_stats(stats_data, df_mrs)
            self._add_project_stats(stats_data, df_mrs)
            self._add_author_stats(stats_data, df_mrs)

            # Créer le DataFrame des statistiques
            stats_df = pd.DataFrame(stats_data, columns=[METRIC_COLUMN_NAME, 'Valeur'])

            # Exporter vers un onglet séparé
            stats_df.to_excel(writer, sheet_name='Statistiques', index=False)

            # Formater l'onglet statistiques
            workbook = writer.book
            stats_worksheet = workbook['Statistiques']
            self._apply_header_style(stats_worksheet, 2)
            self._auto_adjust_columns(stats_worksheet)

        except Exception as e:
            print(f"⚠️ Erreur création statistiques MR: {e}")


# Fonctions utilitaires publiques

    def export_projects(
        self,
        df_projects: pd.DataFrame,
        filename: str = "gitlab_projects.xlsx",
        sheet_name: str = "Gitlab Active Projects"
    ) -> str:
        """
        Exporte les projets vers Excel

        Args:
            df_projects: DataFrame des projets
            filename: Nom du fichier
            sheet_name: Nom de la feuille Excel

        Returns:
            Chemin complet du fichier créé
        """
        if df_projects.empty:
            print("❌ Aucune donnée projet à exporter")
            return ""

        try:
            file_path = self.export_dir / filename
            print(f"📁 Export projets vers: {file_path}")

            # Mapping et nettoyage des colonnes pour les projets
            column_mapping = {
                'id_projet': self.ID_PROJET_HEADER,
                'nom_projet': 'Nom Projet',
                'nom_complet': 'Chemin Complet',
                'namespace': 'Namespace',
                'type_namespace': 'Type Namespace',
                'date_creation': self.DATE_CREATION_HEADER,
                'derniere_activite': 'Date Derniere Activite',
                'dernier_commit': 'Date Dernier Commit',
                'langage_principal': 'Langage Principal',
                'vide': 'Projet Vide'
            }

            # Colonnes à supprimer
            columns_to_drop = ['url_web', 'etat', 'archivé']

            # Supprimer les colonnes non désirées
            df_cleaned = df_projects.copy()
            existing_columns_to_drop = [col for col in columns_to_drop if col in df_cleaned.columns]
            if existing_columns_to_drop:
                df_cleaned = df_cleaned.drop(columns=existing_columns_to_drop)
                print(f"🗑️ Colonnes supprimées: {existing_columns_to_drop}")

            # Renommer les colonnes selon le mapping
            df_renamed = df_cleaned.rename(columns=column_mapping)

            # Trier par ID décroissant (du plus grand au plus petit)
            if self.ID_PROJET_HEADER in df_renamed.columns:
                df_sorted = df_renamed.sort_values(self.ID_PROJET_HEADER, ascending=False)
                print("📊 Projets triés par ID décroissant")
            else:
                df_sorted = df_renamed
                print(f"⚠️ Colonne '{self.ID_PROJET_HEADER}' non trouvée, pas de tri")

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_sorted.to_excel(writer, sheet_name=sheet_name, index=False)

                # Formatage basique
                worksheet = writer.sheets.get(sheet_name)
                if worksheet is not None:
                    self._apply_header_style(worksheet, len(df_sorted.columns))
                    self._auto_adjust_columns(worksheet)

                    # Appliquer l'alignement à gauche pour le contenu
                    self._apply_content_alignment(worksheet)

                    # Ajouter un filtre automatique sur les en-têtes
                    from openpyxl.utils import get_column_letter
                    max_col_letter = get_column_letter(worksheet.max_column)
                    worksheet.auto_filter.ref = f"A1:{max_col_letter}{worksheet.max_row}"

                    worksheet.freeze_panes = "A2"

            print(f"✅ Fichier projets Excel créé: {file_path}")
            print(f"📊 {len(df_sorted)} projets exportés")

            return str(file_path)

        except Exception as e:
            print(f"❌ Erreur lors de l'export projets Excel: {e}")
            return ""

    def export_groups(
        self, df_groups: pd.DataFrame, filename: str = "gitlab_groups.xlsx"
    ) -> str:
        """
        Exporte les groupes vers Excel

        Args:
            df_groups: DataFrame des groupes
            filename: Nom du fichier

        Returns:
            Chemin complet du fichier créé
        """
        if df_groups.empty:
            print("❌ Aucune donnée groupe à exporter")
            return ""

        try:
            file_path = self.export_dir / filename
            print(f"📁 Export groupes vers: {file_path}")

            # Trier les groupes par ID décroissant (plus récent en premier)
            if 'id' in df_groups.columns:
                df_sorted = df_groups.sort_values('id', ascending=False)
                print("🔢 Groupes triés par ID décroissant")
            else:
                df_sorted = df_groups
                print("⚠️ Colonne 'id' non trouvée, pas de tri")

            # Mapping et nettoyage des colonnes pour les groupes
            column_mapping = {
                'id': 'id Groupe',
                'name': 'Nom Groupe',
                'path': 'Chemin Groupe',
                'full_name': 'Nom Complet Groupe',
                'full_path': 'Chemin Complet Groupe',
                'created_at': self.DATE_CREATION_HEADER,
                'parent_id': 'id Parent',
                'parent_name': 'Groupe Parent',
                'projects_count': 'Nombre Projets',
                'members_count': 'Nombre Membres',
                'subgroups_count': 'Nombre Sous-Groupes'
            }

            # Colonnes à supprimer pour les groupes
            columns_to_drop = ['description', 'visibility', 'web_url']

            # Supprimer les colonnes non désirées
            df_cleaned = df_sorted.copy()
            existing_columns_to_drop = [col for col in columns_to_drop if col in df_cleaned.columns]
            if existing_columns_to_drop:
                df_cleaned = df_cleaned.drop(columns=existing_columns_to_drop)
                print(f"🗑️ Colonnes supprimées: {existing_columns_to_drop}")

            # Renommer les colonnes selon le mapping
            df_renamed = df_cleaned.rename(columns=column_mapping)

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_renamed.to_excel(writer, sheet_name='Gitlab Groups', index=False)

                # Formatage basique
                worksheet = writer.sheets.get('Gitlab Groups')
                if worksheet is not None:
                    self._apply_header_style(worksheet, len(df_renamed.columns))
                    self._auto_adjust_columns(worksheet)

                    # Appliquer l'alignement à gauche pour le contenu
                    self._apply_content_alignment(worksheet)

                    # Ajouter un filtre automatique sur les en-têtes
                    from openpyxl.utils import get_column_letter
                    max_col_letter = get_column_letter(worksheet.max_column)
                    worksheet.auto_filter.ref = f"A1:{max_col_letter}{worksheet.max_row}"

                    worksheet.freeze_panes = "A2"

            print(f"✅ Fichier groupes Excel créé: {file_path}")
            print(f"📊 {len(df_renamed)} groupes exportés")

            return str(file_path)

        except Exception as e:
            print(f"❌ Erreur lors de l'export groupes Excel: {e}")
            return ""

    def export_combined_report(self, df_users: pd.DataFrame, df_projects: pd.DataFrame,
                             stats: Dict[str, Any], filename: str = "gitlab_rapport_complet.xlsx") -> str:
        """
        Exporte un rapport combiné avec utilisateurs, projets et statistiques

        Args:
            df_users: DataFrame des utilisateurs
            df_projects: DataFrame des projets
            stats: Dictionnaire des statistiques
            filename: Nom du fichier

        Returns:
            Chemin complet du fichier créé
        """
        try:
            file_path = self.export_dir / filename
            print(f"📁 Export rapport complet vers: {file_path}")

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Feuille utilisateurs
                if not df_users.empty:
                    df_users.to_excel(writer, sheet_name='Utilisateurs', index=False)
                    worksheet = writer.sheets.get('Utilisateurs')
                    if worksheet is not None:
                        self._apply_header_style(worksheet, len(df_users.columns))
                        self._auto_adjust_columns(worksheet)
                        worksheet.freeze_panes = "A2"

                # Feuille projets
                if not df_projects.empty:
                    df_projects.to_excel(writer, sheet_name='Projets', index=False)
                    worksheet = writer.sheets.get('Projets')
                    if worksheet is not None:
                        self._apply_header_style(worksheet, len(df_projects.columns))
                        self._auto_adjust_columns(worksheet)
                        worksheet.freeze_panes = "A2"

                # Feuille statistiques
                if stats:
                    stats_df = pd.DataFrame(list(stats.items()), columns=[METRIC_COLUMN_NAME, 'Valeur'])
                    stats_df.to_excel(writer, sheet_name='Statistiques', index=False)
                    worksheet = writer.sheets.get('Statistiques')
                    if worksheet is not None:
                        self._apply_header_style(worksheet, 2)
                        self._auto_adjust_columns(worksheet)

            print(f"✅ Rapport complet Excel créé: {file_path}")
            return str(file_path)

        except Exception as e:
            print(f"❌ Erreur lors de l'export rapport complet: {e}")
            return ""


def export_projects_to_excel(
    df_projects: pd.DataFrame, filename: str = "gitlab_projects.xlsx"
) -> str:
    """
    Fonction simple pour exporter les projets vers Excel

    Args:
        df_projects: DataFrame des projets
        filename: Nom du fichier

    Returns:
        Chemin du fichier créé
    """
    exporter = GitLabExcelExporter()
    return exporter.export_projects(df_projects, filename)


def export_merge_requests_to_excel(
    df_mrs: pd.DataFrame, filename: str = "gitlab_merge_requests.xlsx"
) -> Optional[str]:
    """
    Fonction utilitaire pour exporter les Merge Requests GitLab vers Excel

    Args:
        df_mrs: DataFrame des Merge Requests
        filename: Nom du fichier de sortie

    Returns:
        Chemin du fichier créé ou None si erreur
    """
    exporter = GitLabExcelExporter()
    return exporter.export_merge_requests(df_mrs, filename)


def export_events_to_excel(df_events: pd.DataFrame, filename: str = "gitlab_events.xlsx") -> str:
    """
    Fonction utilitaire pour exporter les événements GitLab vers Excel

    Args:
        df_events: DataFrame avec les données d'événements
        filename: Nom du fichier

    Returns:
        Chemin du fichier créé
    """
    exporter = GitLabExcelExporter()
    return exporter.export_events(df_events, filename)


def export_groups_to_excel(df_groups: pd.DataFrame, filename: str = "gitlab_groups.xlsx") -> str:
    """
    Fonction utilitaire pour exporter les groupes GitLab vers Excel

    Args:
        df_groups: DataFrame avec les données de groupes
        filename: Nom du fichier

    Returns:
        Chemin du fichier créé
    """
    exporter = GitLabExcelExporter()
    return exporter.export_groups(df_groups, filename)


def _apply_basic_formatting(ws) -> None:
    """Applique un formatage de base à la feuille Excel"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Appliquer le style aux en-têtes
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _get_column_letter(column) -> Optional[str]:
    """Récupère la lettre de colonne du premier élément non-fusionné"""
    for cell in column:
        if hasattr(cell, 'column_letter'):
            return cell.column_letter
    return None


def _auto_adjust_columns(ws) -> None:
    """Auto-ajuste la largeur des colonnes"""
    for column in ws.columns:
        column_letter = _get_column_letter(column)
        if column_letter:
            # Calcul simple de la largeur
            max_length = 0
            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_value))
                except Exception:
                    pass
            # Largeur finale avec marge et limites
            adjusted_width = min(max(max_length + 3, 12), 60)
            ws.column_dimensions[column_letter].width = adjusted_width


def export_to_excel(df: pd.DataFrame, filename: str, sheet_name: str = "Données",
                    subfolder: str = "gitlab") -> str:
    """
    Fonction générique d'export vers Excel

    Args:
        df: DataFrame à exporter
        filename: Nom du fichier (sans extension)
        sheet_name: Nom de l'onglet
        subfolder: Sous-dossier dans exports/

    Returns:
        Chemin du fichier créé
    """
    if df.empty:
        raise ValueError("DataFrame vide - aucune donnée à exporter")

    # Déterminer le répertoire d'export (remonter 4 niveaux depuis kenobi_tools/gitlab/exporters/)
    current_dir = Path(__file__).parent.parent.parent.parent
    export_dir = current_dir / "exports" / subfolder
    export_dir.mkdir(parents=True, exist_ok=True)

    # Ajouter extension si manquante
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'

    # Chemin complet du fichier
    file_path = export_dir / filename

    try:
        # Créer le workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = sheet_name

            # Ajouter les données
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # Appliquer le formatage
            _apply_basic_formatting(ws)
            _auto_adjust_columns(ws)

        # Sauvegarder
        wb.save(file_path)
        return str(file_path)

    except Exception as e:
        raise RuntimeError(f"Erreur lors de l'export: {e}") from e


if __name__ == "__main__":
    """Test de l'exporteur Excel"""
    from pathlib import Path

    # Test avec des données fictives
    print("🧪 Test de l'exporteur Excel GitLab")
    print("=" * 50)

    # Créer des données de test
    test_users = pd.DataFrame({
        'id_utilisateur': [1, 2, 3],
        'nom_utilisateur': ['user1', 'user2', 'user3'],
        'email': ['user1@test.com', 'user2@test.com', 'user3@test.com'],
        'nom_complet': ['Utilisateur Un', 'Utilisateur Deux', 'Utilisateur Trois'],
        'admin': ['Non', 'Oui', 'Non'],
        'etat': ['Active', 'Active', 'Blocked'],
        'type_utilisateur': ['Humain', 'Humain', 'Humain'],
        'date_creation': ['01/01/2024 10:00:00', '02/01/2024 11:00:00', '03/01/2024 12:00:00']
    })

    # Test d'export
    exporter = GitLabExcelExporter()
    file_path = exporter.export_users(test_users, "test_users.xlsx")

    if file_path:
        print(f"✅ Test réussi! Fichier créé: {file_path}")
    else:
        print("❌ Test échoué")
