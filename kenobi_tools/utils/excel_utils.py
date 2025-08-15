#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Utilitaires pour l'export Excel standardisé
Module centralisé pour uniformiser le formatage des fichiers Excel
"""

import os
from typing import Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def apply_standard_excel_formatting(
    worksheet,
    enable_autofilter: bool = True,
    freeze_first_row: bool = True,
    header_color: str = "366092",
    header_font_color: str = "FFFFFF",
    content_alignment: str = "left"
):
    """
    Applique le formatage standard aux feuilles Excel
    """
    if not worksheet:
        return
    
    try:
        # Activer le filtre automatique
        if enable_autofilter and worksheet.max_row > 1:
            max_col_letter = get_column_letter(worksheet.max_column)
            worksheet.auto_filter.ref = f"A1:{max_col_letter}{worksheet.max_row}"
    except Exception as e:
        print(f"⚠️ Erreur filtre automatique: {e}")
    
    try:
        # Figer la première ligne
        if freeze_first_row:
            worksheet.freeze_panes = "A2"
    except Exception as e:
        print(f"⚠️ Erreur gel première ligne: {e}")
    
    try:
        # Formater les en-têtes (première ligne)
        if worksheet.max_row > 0:
            header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            header_font = Font(bold=True, color=header_font_color)
            header_alignment = Alignment(horizontal="left", vertical="center")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
    except Exception as e:
        print(f"⚠️ Erreur formatage en-têtes: {e}")
    
    try:
        # Formater le contenu (toutes les autres lignes)
        if worksheet.max_row > 1:
            content_alignment_obj = Alignment(horizontal=content_alignment, vertical="center")
            for row in range(2, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = content_alignment_obj
    except Exception as e:
        print(f"⚠️ Erreur alignement contenu: {e}")
    
    try:
        # Ajuster automatiquement la largeur des colonnes
        auto_adjust_column_widths(worksheet)
    except Exception as e:
        print(f"⚠️ Erreur ajustement colonnes: {e}")


def auto_adjust_column_widths(worksheet, max_width: int = 50, min_width: int = 10):
    """
    Ajuste automatiquement la largeur des colonnes
    
    Args:
        worksheet: Feuille Excel
        max_width: Largeur maximum des colonnes
        min_width: Largeur minimum des colonnes
    """
    if not worksheet:
        return
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        
        # Calculer la largeur ajustée
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def export_dataframe_to_excel_light(
    df: pd.DataFrame,
    filename: str,
    sheet_name: str = "Data",
    column_mapping: Optional[Dict[str, str]] = None,
    exports_dir: str = "exports/gitlab",
    auto_adjust_columns: bool = True
) -> str:
    """
    Exporte un DataFrame vers Excel avec formatage minimal (optimisé pour gros volumes)
    
    Args:
        df: DataFrame à exporter
        filename: Nom du fichier (avec ou sans extension)
        sheet_name: Nom de la feuille Excel
        column_mapping: Mapping pour renommer les colonnes
        exports_dir: Répertoire de destination
        auto_adjust_columns: Ajuster automatiquement la largeur des colonnes
    
    Returns:
        Chemin complet du fichier généré
    """
    # Préparer le nom de fichier
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    # Créer le répertoire exports s'il n'existe pas
    os.makedirs(exports_dir, exist_ok=True)
    filepath = os.path.join(exports_dir, filename)
    
    # Préparer le DataFrame
    df_to_export = df.copy()
    
    # Appliquer le mapping des colonnes si fourni
    if column_mapping:
        # Filtrer seulement les colonnes qui existent
        available_columns = [col for col in column_mapping.keys() if col in df.columns]
        df_to_export = df_to_export[available_columns]
        df_to_export = df_to_export.rename(columns=column_mapping)
    
    # Export rapide avec formatage minimal
    try:
        print(f"📝 Export Excel en cours ({len(df_to_export)} lignes, {len(df_to_export.columns)} colonnes)...")
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df_to_export.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Formatage minimal pour améliorer la lisibilité
            worksheet = writer.sheets.get(sheet_name)
            if worksheet and auto_adjust_columns:
                # Seulement l'ajustement des colonnes (rapide)
                auto_adjust_column_widths_light(worksheet)
        
        # Vérifier que le fichier a été créé
        if os.path.exists(filepath) and os.path.getsize(filepath) > 0:
            size_mb = os.path.getsize(filepath) / 1024 / 1024
            print(f"✅ Fichier Excel créé: {size_mb:.1f} MB")
        else:
            print(f"❌ Erreur: fichier Excel vide ou non créé")
            
    except Exception as e:
        print(f"❌ Erreur lors de la création du fichier Excel: {e}")
        raise
    
    return filepath


def auto_adjust_column_widths_light(worksheet, max_width: int = 50, min_width: int = 8):
    """
    Ajuste automatiquement la largeur des colonnes (version optimisée)
    
    Args:
        worksheet: Feuille Excel
        max_width: Largeur maximum des colonnes
        min_width: Largeur minimum des colonnes
    """
    if not worksheet:
        return
    
    # Optimisation: ne regarder que les 100 premières lignes pour calculer la largeur
    max_rows_to_check = min(100, worksheet.max_row)
    
    for col_num in range(1, worksheet.max_column + 1):
        max_length = min_width
        column_letter = worksheet.cell(row=1, column=col_num).column_letter
        
        # Vérifier seulement un échantillon des cellules
        for row_num in range(1, max_rows_to_check + 1):
            try:
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    length = len(str(cell_value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        
        # Calculer la largeur ajustée
        adjusted_width = min(max(max_length + 1, min_width), max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def get_standard_formatting_options() -> Dict[str, Any]:
    """
    Retourne les options de formatage standard pour tous les exports
    
    Returns:
        Dictionnaire des options de formatage par défaut
    """
    return {
        'enable_autofilter': True,
        'freeze_first_row': True,
        'header_color': "366092",  # Bleu professionnel
        'header_font_color': "FFFFFF"  # Blanc
    }


# Constantes pour l'uniformisation
STANDARD_EXPORTS_DIR = "exports/gitlab"  # Répertoire spécifique GitLab
STANDARD_HEADER_COLOR = "366092"
STANDARD_HEADER_FONT_COLOR = "FFFFFF"
STANDARD_CONTENT_ALIGNMENT = "left"  # Alignement à gauche par défaut
